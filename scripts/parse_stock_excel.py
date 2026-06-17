import sys
import json
from pathlib import Path
import openpyxl

FILE = sys.argv[1] if len(sys.argv) > 1 else r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\Stock Management 2025 final.xlsx"

wb = openpyxl.load_workbook(FILE, data_only=True)
summary = {
    "file": str(FILE),
    "sheets": []
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_info = {
        "name": sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "headers": [],
        "rows_count": 0,
        "sample_rows": [],
        "columns": {}
    }

    # Collect rows and detect header: first non-empty row
    rows = list(ws.iter_rows(values_only=True))
    # find header row index
    header_idx = None
    for i, row in enumerate(rows):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header_idx = i
            break

    if header_idx is None:
        summary["sheets"].append(sheet_info)
        continue

    headers = [str(c).strip() if c is not None else f"col_{i+1}" for i, c in enumerate(rows[header_idx])]
    sheet_info['headers'] = headers

    data_rows = rows[header_idx+1:]
    # filter out fully empty rows
    filtered = [r for r in data_rows if not all(cell is None or str(cell).strip()=="" for cell in r)]
    sheet_info['rows_count'] = len(filtered)

    # Column analysis
    cols = {h: {"non_empty":0, "types":{}, "sample_values":[]} for h in headers}
    for r in filtered:
        for i, h in enumerate(headers):
            val = r[i] if i < len(r) else None
            if val is not None and str(val).strip() != "":
                cols[h]['non_empty'] += 1
                t = type(val).__name__
                cols[h]['types'][t] = cols[h]['types'].get(t, 0) + 1
                if len(cols[h]['sample_values']) < 5:
                    cols[h]['sample_values'].append(val)

    sheet_info['columns'] = cols

    # sample first 10 rows
    for r in filtered[:10]:
        row_data = [v for v in r]
        sheet_info['sample_rows'].append(row_data)

    summary['sheets'].append(sheet_info)

# write JSON
out = Path(__file__).parent / 'stock_excel_summary.json'
with out.open('w', encoding='utf-8') as f:
    json.dump(summary, f, indent=2, default=str)

print(f"Summary written to: {out}")
print(json.dumps({"file": summary['file'], "sheets_count": len(summary['sheets'])}))
