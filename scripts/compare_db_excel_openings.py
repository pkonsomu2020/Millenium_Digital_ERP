import json
import sys
from pathlib import Path
import openpyxl

EXCEL = sys.argv[1] if len(sys.argv)>1 else r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\Stock Management 2025 final.xlsx"
DB_JSON = sys.argv[2] if len(sys.argv)>2 else r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\backend\scripts\stock_seed_007.json"
MONTH_KEY = sys.argv[3] if len(sys.argv)>3 else '2025-09'  # default to September 2025
CATEGORIES = ['Kitchen Essentials', 'Washroom Essentials', 'Snacks']

wb = openpyxl.load_workbook(EXCEL, data_only=True)
with open(DB_JSON, 'r', encoding='utf-8') as f:
    db = json.load(f)

# Build DB opening map: (category, item_name) -> p1_opening for given month_key
db_openings = {}
for e in db.get('entries', []):
    if e.get('month_key') != MONTH_KEY:
        continue
    cat = e.get('category')
    item = e.get('item_name')
    p1_opening = e.get('p1_opening')
    db_openings.setdefault(cat, {})[item.strip()] = p1_opening

# From Excel, find sheet for month (sheet named 'Sep' or startswith 'Sep')
# The workbook has sheet 'Sep' per earlier inspection.
possible_sheet_names = [s for s in wb.sheetnames if s.lower().startswith('sep') or 'sept' in s.lower() or s.lower().startswith('sep')]
if not possible_sheet_names:
    # fallback: look for sheet titled 'Sep' exact
    possible_sheet_names = ['Sep'] if 'Sep' in wb.sheetnames else []

if not possible_sheet_names:
    print('No sheet for September found in workbook. Available sheets:', wb.sheetnames)
    sys.exit(1)

sheet = wb[possible_sheet_names[0]]
rows = list(sheet.iter_rows(values_only=True))

# find header row index with "ITEM NAME"
header_idx = None
for i, row in enumerate(rows):
    if any(cell and isinstance(cell, str) and 'ITEM NAME' in cell.upper() for cell in row if cell is not None):
        header_idx = i
        break

if header_idx is None:
    print('Header row not found in sheet', sheet.title)
    sys.exit(1)

headers = [ (cell if cell is not None else '') for cell in rows[header_idx] ]
# Determine columns: find index of 'ITEM NAME', then the first 'Opening' under period1 is at index after item_name
item_col = None
for idx, h in enumerate(headers):
    if isinstance(h, str) and 'ITEM' in h.upper():
        item_col = idx
        break

# For period1 opening, find the first 'Opening' column after item_col
p1_open_col = None
for idx in range(item_col+1, len(headers)):
    h = headers[idx]
    if isinstance(h, str) and 'OPEN' in h.upper():
        p1_open_col = idx
        break

if item_col is None or p1_open_col is None:
    print('Could not determine item or period1 opening columns')
    print('Headers:', headers)
    sys.exit(1)

# Parse rows after header until end; categories indicated by rows with 'KITCHEN' etc.
excel_openings = {c: {} for c in CATEGORIES}
current_category = None
for r in rows[header_idx+1:]:
    if not any(cell is not None and str(cell).strip()!='' for cell in r):
        continue
    first = r[item_col]
    if isinstance(first, str) and 'KITCHEN ESSENTIALS' in first.upper():
        current_category = 'Kitchen Essentials'
        continue
    if isinstance(first, str) and 'WASHROOM ESSENTIALS' in first.upper():
        current_category = 'Washroom Essentials'
        continue
    if isinstance(first, str) and 'SNACKS' in first.upper():
        current_category = 'Snacks'
        continue
    if current_category in CATEGORIES:
        item_name = r[item_col]
        if item_name is None:
            continue
        # Normalize item name (strip units in parentheses)
        name = str(item_name).strip()
        # p1 opening may be int/float or empty
        val = None
        try:
            val = r[p1_open_col]
            if val is None:
                # try p1 opening from two cols earlier/following if formatting differs
                pass
        except IndexError:
            val = None
        # store
        excel_openings[current_category][name] = val if val is not None else 0

# Now compare DB openings vs Excel openings
report = []
for cat in CATEGORIES:
    db_items = db_openings.get(cat, {})
    excel_items = excel_openings.get(cat, {})
    all_items = sorted(set(list(db_items.keys()) + list(excel_items.keys())))
    diffs = []
    for it in all_items:
        db_val = db_items.get(it)
        ex_val = excel_items.get(it)
        # Normalize numbers
        db_v = db_val if db_val is not None else 0
        ex_v = ex_val if ex_val is not None else 0
        try:
            db_num = float(db_v)
        except:
            db_num = db_v
        try:
            ex_num = float(ex_v)
        except:
            ex_num = ex_v
        if db_num != ex_num:
            diffs.append({'item': it, 'db_opening': db_v, 'excel_opening': ex_v})
    report.append({'category': cat, 'differences': diffs, 'db_count': len(db_items), 'excel_count': len(excel_items)})

out = Path(__file__).parent / f'openings_compare_{MONTH_KEY}.json'
with out.open('w', encoding='utf-8') as f:
    json.dump(report, f, indent=2, default=str)

print('Comparison written to', out)
print(json.dumps(report, indent=2))
