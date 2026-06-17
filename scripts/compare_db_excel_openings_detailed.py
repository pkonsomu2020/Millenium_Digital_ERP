import json
import sys
from pathlib import Path
import openpyxl

EXCEL = sys.argv[1] if len(sys.argv)>1 else r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\Stock Management 2025 final.xlsx"
DB_JSON = sys.argv[2] if len(sys.argv)>2 else r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\backend\scripts\stock_seed_007.json"
MONTH_KEY = sys.argv[3] if len(sys.argv)>3 else '2025-09'  # default to September 2025
CATEGORIES = ['Kitchen Essentials', 'Washroom Essentials', 'Snacks']

wb = openpyxl.load_workbook(EXCEL, data_only=True)
with open(DB_JSON, 'r', encoding='utf-8') as f:
    db = json.load(f)

# Build DB opening map: (category, item_name) -> p1_opening for given month_key
db_openings = {}
for e in db.get('entries', []):
    if e.get('month_key') != MONTH_KEY:
        continue
    cat = e.get('category')
    item = e.get('item_name')
    p1_opening = e.get('p1_opening')
    db_openings.setdefault(cat, {})[item.strip()] = p1_opening

# Excel sheet selection for September
sheet_name = None
for s in wb.sheetnames:
    if s.lower().startswith('sep'):
        sheet_name = s
        break
if not sheet_name:
    print('September sheet not found')
    sys.exit(1)

sheet = wb[sheet_name]
rows = list(sheet.iter_rows(values_only=True))
# find header
header_idx = None
for i, row in enumerate(rows):
    if any(cell and isinstance(cell, str) and 'ITEM NAME' in cell.upper() for cell in row if cell is not None):
        header_idx = i
        break
if header_idx is None:
    print('Header not found')
    sys.exit(1)
headers = [cell if cell is not None else '' for cell in rows[header_idx]]
# locate item and p1 opening indices
item_col = None
for idx, h in enumerate(headers):
    if isinstance(h, str) and 'ITEM' in h.upper():
        item_col = idx
        break
# find the 'Opening' header that belongs to PERIOD 1 by looking at column groups: the sheet seems to have PERIOD1 (cols 2-5), PERIOD2 (6-9), then monthly summary 10-13
# Identify by finding first 'Opening' after item_col
p1_open_col = None
for idx in range(item_col+1, len(headers)):
    h = headers[idx]
    if isinstance(h, str) and 'OPEN' in h.upper():
        p1_open_col = idx
        break

excel_openings = {c: {} for c in CATEGORIES}
current_category = None
for r in rows[header_idx+1:]:
    if not any(cell is not None and str(cell).strip()!='' for cell in r):
        continue
    first = r[item_col]
    if isinstance(first, str) and 'KITCHEN' in first.upper():
        current_category = 'Kitchen Essentials'; continue
    if isinstance(first, str) and 'WASHROOM' in first.upper():
        current_category = 'Washroom Essentials'; continue
    if isinstance(first, str) and 'SNACKS' in first.upper():
        current_category = 'Snacks'; continue
    if current_category in CATEGORIES:
        item = r[item_col]
        if item is None: continue
        name = str(item).strip()
        val = None
        if p1_open_col < len(r):
            val = r[p1_open_col]
        excel_openings[current_category][name] = val if val is not None else 0

report = {}
for cat in CATEGORIES:
    db_items = db_openings.get(cat, {})
    excel_items = excel_openings.get(cat, {})
    only_in_db = sorted([i for i in db_items.keys() if i not in excel_items.keys()])
    only_in_excel = sorted([i for i in excel_items.keys() if i not in db_items.keys()])
    mismatches = []
    for it in set(db_items.keys()).intersection(excel_items.keys()):
        db_v = db_items.get(it) or 0
        ex_v = excel_items.get(it) or 0
        try:
            if float(db_v) != float(ex_v):
                mismatches.append({'item': it, 'db_opening': db_v, 'excel_opening': ex_v})
        except Exception:
            if db_v != ex_v:
                mismatches.append({'item': it, 'db_opening': db_v, 'excel_opening': ex_v})
    report[cat] = {'only_in_db': only_in_db, 'only_in_excel': only_in_excel, 'mismatches': mismatches, 'db_count': len(db_items), 'excel_count': len(excel_items)}

out = Path(__file__).parent / f'openings_detailed_{MONTH_KEY}.json'
with out.open('w', encoding='utf-8') as f:
    json.dump(report, f, indent=2, default=str)

print('Wrote detailed comparison to', out)
print(json.dumps(report, indent=2))
