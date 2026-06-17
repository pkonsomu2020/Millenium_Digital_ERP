import csv
import json
import re
from pathlib import Path
import openpyxl
import sys

CSV_PATH = Path(r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\CSV DATA\DATABASE\stock_items_rows.csv")
EXCEL_PATH = Path(r"D:\backup\Downloads\ALX PROJECTS\Millenium_Digital_ERP\Stock Management 2025 final.xlsx")
CATEGORIES = ['Kitchen Essentials', 'Washroom Essentials', 'Snacks']

def normalize_name(name):
    if name is None:
        return ''
    s = str(name).strip()
    # remove unit parentheses like 'Sugar (kg)'
    s = re.sub(r"\s*\(.*\)$", '', s).strip()
    return s

# Load CSV
csv_items = {}
with CSV_PATH.open(encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for r in reader:
        cat = r['category'].strip()
        name = normalize_name(r['item_name'])
        try:
            qty = float(r['current_quantity'])
        except:
            qty = None
        csv_items.setdefault(cat, {})[name] = {'raw_name': r['item_name'], 'qty': qty, 'unit': r.get('unit')}

# Parse Excel: iterate sheets and capture latest Final Closing per item
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
excel_latest = {c: {} for c in CATEGORIES}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    # find header row index where 'ITEM NAME' appears
    header_idx = None
    for i, row in enumerate(rows):
        if any(isinstance(cell, str) and 'ITEM NAME' in cell.upper() for cell in row if cell is not None):
            header_idx = i
            break
    if header_idx is None:
        continue
    headers = [cell if cell is not None else '' for cell in rows[header_idx]]
    # find item col
    item_col = None
    for idx, h in enumerate(headers):
        if isinstance(h, str) and 'ITEM' in h.upper():
            item_col = idx
            break
    # find final closing column (header contains 'FINAL')
    final_col = None
    for idx, h in enumerate(headers):
        if isinstance(h, str) and 'FINAL' in h.upper():
            final_col = idx
            break
    if item_col is None or final_col is None:
        continue

    current_category = None
    for r in rows[header_idx+1:]:
        if not any(cell is not None and str(cell).strip()!='' for cell in r):
            continue
        first = r[item_col]
        if isinstance(first, str) and 'KITCHEN ESSENTIALS' in first.upper():
            current_category = 'Kitchen Essentials'; continue
        if isinstance(first, str) and 'WASHROOM ESSENTIALS' in first.upper():
            current_category = 'Washroom Essentials'; continue
        if isinstance(first, str) and 'SNACKS' in first.upper():
            current_category = 'Snacks'; continue
        if current_category in CATEGORIES:
            item = r[item_col]
            if item is None:
                continue
            name = normalize_name(item)
            val = None
            try:
                val = r[final_col]
            except IndexError:
                val = None
            if val is None:
                # treat empty as 0
                try:
                    val_num = 0
                except:
                    val_num = 0
            else:
                try:
                    val_num = float(val)
                except:
                    val_num = val
            # update latest value (overwrites earlier month values as we iterate chronologically)
            excel_latest[current_category][name] = val_num

# Compare for each category
report = {}
for cat in CATEGORIES:
    csv_map = csv_items.get(cat, {})
    excel_map = excel_latest.get(cat, {})
    only_in_csv = [n for n in sorted(csv_map.keys()) if n not in excel_map]
    only_in_excel = [n for n in sorted(excel_map.keys()) if n not in csv_map]
    diffs = []
    for name in sorted(set(csv_map.keys()).intersection(excel_map.keys())):
        csv_q = csv_map[name]['qty']
        excel_q = excel_map[name]
        csv_qn = csv_q if csv_q is not None else 0
        excel_qn = excel_q if excel_q is not None else 0
        # compare numeric
        try:
            if float(csv_qn) != float(excel_qn):
                diffs.append({'item': name, 'csv_qty': csv_qn, 'excel_latest_final_closing': excel_qn, 'csv_raw_name': csv_map[name]['raw_name']})
        except Exception:
            if csv_qn != excel_qn:
                diffs.append({'item': name, 'csv_qty': csv_qn, 'excel_latest_final_closing': excel_qn, 'csv_raw_name': csv_map[name]['raw_name']})
    report[cat] = {'only_in_csv': only_in_csv, 'only_in_excel': only_in_excel, 'differences': diffs, 'csv_count': len(csv_map), 'excel_count': len(excel_map)}

out = Path('scripts') / 'live_csv_vs_excel_latest.json'
with out.open('w', encoding='utf-8') as f:
    json.dump(report, f, indent=2)

print('Wrote comparison to', out)
print(json.dumps(report, indent=2))
